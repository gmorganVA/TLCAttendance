"""Report event attendance in Trail Life Connect"""
import argparse
import logging
import configparser
import getpass
import time
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook

##################################################################################
class TLCAttendance:

    ##############################################################################
    def __init__(self):
        """Read command line and config file input and setup logging"""
        # Process command line input
        parser = argparse.ArgumentParser(
            prog="TLCAttendance",
            description="Report event attendance in Trail Life Connect"
        )
        parser.add_argument('-f', '--filename', help='Attendance file', required=True)
        parser.add_argument('-e', '--event', help='Event name', required=True)
        parser.add_argument('-l', '--log', help='Log file',
            default='./TLCAttendance.log')
        parser.add_argument('-c', '--config', help='Config file',
            default='./.config.ini')
        self.args = parser.parse_args()
        # Set up logging
        logging.basicConfig(level=logging.INFO, filename=self.args.log)
        logging.info("Using attendance file: %s", self.args.filename)
        logging.info("Using log file: %s", self.args.log)
        logging.info("Using config file: %s", self.args.config)
        self.ID_LESSON_PLAN_CBOX = 'use-lesson-plans'
        self.ID_EVENT_SELECTION = 'select2-event-id-container'
        self.CLS_EVENT_SEARCH = 'select2-search__field'
        self.driver = None

    ##############################################################################
    def __del__(self):
        if self.driver is not None:
            self.driver.quit()

    ##############################################################################
    def readConfigItem(self, section, fieldName, promptStr, defaultValue):
        # Add section if it doesn't exist in config yet
        if section not in self.config:
            self.config[section] = {}
        # Read config value from user if not in config file
        if not fieldName in self.config[section]:
            logging.info(f"Config file missing [{section}] {fieldName}")
            value = input(promptStr) or None
            # Use default if user doesn't provide an input
            if value is None:
                value = defaultValue
                logging.info(f"No value provided for [{section}] {fieldName}, using {defaultValue}")
            # Set new value in config file
            self.config[section][fieldName] = value
            self.configChanged = True
        setattr(self, fieldName, self.config[section][fieldName])

    ##############################################################################
    def readConfig(self):
        self.configChanged = False
        self.config = configparser.ConfigParser()
        self.config.read(self.args.config)

        # Read values from the default section of the config file
        self.readConfigItem('DEFAULT', 'server', 'Trail Life Connect website: ',
            'www.traillifeconnect.com')
        self.setURLs()
        self.readConfigItem('DEFAULT', 'browser',
            'Web browser to use (Firefox, Chrome, Edge): ', 'Firefox')
        self.readConfigItem('DEFAULT', 'sheetName', 'Attendance sheet name: ',
            'Attendance')

        # Read values from the server specific section
        self.readConfigItem(self.server, 'email', 'Trail Life Connect username: ',
            'user@example.com')
        self.readConfigItem(self.server, 'initial_wait_secs', 
            'Time to wait for website in seconds: ', '30')

        if self.configChanged:
            logging.info("Writing changes to config file %s", self.args.config)
            with open(self.args.config, 'w') as configFile:
                self.config.write(configFile)

    ##############################################################################
    def setURLs(self):
        self.BASE_URL = f"https://{self.server}/"
        self.DASHBOARD_URL = f"{self.BASE_URL}dashboard"
        self.LOGIN_URL = f"{self.BASE_URL}login"

    ##############################################################################
    def getDriver(self):
        """Get Selenium web driver based on selected browser"""
        if self.browser == 'Firefox':
            self.driver = webdriver.Firefox()
        elif self.browser == 'Chrome':
            self.driver = webdriver.Chrome()
        elif self.browser == 'Edge':
            self.driver = webdriver.Edge()
        else: # Default case: Use Firefox
            self.driver = webdriver.Firefox()
    
    ##############################################################################
    def scrollFirefox(self, element):
        self.driver.execute_script("arguments[0].scrollIntoView()", element)

    ##############################################################################
    def login(self):
        """Login to Trail Life Connect"""
        print(f"Username: {self.email}")
        tlcPass = getpass.getpass(f"Password for {self.server}: ")
        self.getDriver()
        self.driver.get(self.BASE_URL)
        wait = WebDriverWait(self.driver, self.initial_wait_secs)
        loginBtn = wait.until(EC.element_to_be_clickable((By.NAME, "login-button")))
        emailField = self.driver.find_element(By.ID, "loginform-email")
        emailField.send_keys(self.email)
        loginComplete = False
        while not loginComplete:
            passField = self.driver.find_element(By.ID, "loginform-password")
            passField.send_keys(tlcPass)
            loginBtn.click()
            waitLogin = WebDriverWait(self.driver, 10)
            try:
                waitLogin.until(EC.any_of(
                    EC.visibility_of_element_located((By.CLASS_NAME, "help-block-error")),
                    EC.url_to_be(self.DASHBOARD_URL)))
            except TimeoutException:
                print(f'ERROR: Login failed {self.driver.current_url}')
                return False    

            if (self.driver.current_url == self.DASHBOARD_URL):
                loginComplete = True
            else:
                print(f'ERROR: Failed to load dashboard after login {self.driver.current_url}')
                return False
            
        return True

    ##############################################################################
    def logout(self):
        """Logout from Trail Life Connect"""
        self.driver.get(f"https://{self.server}/logout")
        self.driver.close()
    
    ##############################################################################
    def loadAttendanceData(self):
        """Read excel worksheet from barcode reader software"""
        self.attendees = []
        workBook = load_workbook(self.args.filename)
        if self.sheetName in workBook.sheetnames:
            attendanceSheet = workBook[self.sheetName]
        
            for col in attendanceSheet['A']:
                self.attendees.append(col.value)
            workBook.close()
        else:
            print(f'ERROR: Sheet {self.sheetName} not found in {self.args.filename}')
        return len(self.attendees)

    ##############################################################################
    def recordAttendance(self):
        """Record event attendance"""
        self.driver.get(f"https://{self.server}/attendance")
        # Pick event that attendance is being tracked for
        time.sleep(2)
        
        print(f'Select event: {self.args.event}')
        eventIdSelectElement = self.driver.find_element(By.ID, "event-id")
        #eventIdSelectElement.click()
        eventIdSelect = Select(eventIdSelectElement)
        eventIdClickable = self.driver.find_element(By.ID, 'select2-event-id-container')
        eventActions = ActionChains(self.driver)
        eventActions.click(eventIdClickable)
        eventActions.send_keys(self.args.event)
        eventActions.pause(1)
        eventActions.send_keys(Keys.ENTER)
        eventActions.perform()
        selectedEventValue = eventIdSelect.first_selected_option.get_attribute("value")
        print(f'Selected option code: {selectedEventValue}')
        time.sleep(1)

        print('Check use lesson plan')
        useLessonPlansInput = self.driver.find_element(By.ID, self.ID_LESSON_PLAN_CBOX)
        #waitLessonPlan = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable(useLessonPlansCheckbox))
        lessonPlanChecked = useLessonPlansInput.get_attribute('value')
        print(f'Lesson plan checked: {lessonPlanChecked}')
        if lessonPlanChecked == '0':
            useLessonPlansInput.find_element(By.XPATH, '..').click()
            time.sleep(1)
        
        # Check each attendee's name
        attendeeCount = 0
        for attendee in self.attendees:
            attendeeCount += 1
            print(f'Check attendee: {attendee}')
            endStr = 'attended'
            attendeeCboxId = f'{attendee}-{selectedEventValue}-{endStr}'
            attendeeXPath = f"//input[@id = '{attendeeCboxId}']"
            print(f'Find attendee by xpath: {attendeeXPath}')
            attendeeCboxInput = self.driver.find_element(By.XPATH, attendeeXPath)
            attendeeCboxParent = attendeeCboxInput.find_element(By.XPATH, '..')

            print('Before move to element')
            attendeeCboxChecked = attendeeCboxInput.get_attribute('value')
            print(f'Attendee checkbox value: {attendeeCboxChecked}')
            if attendeeCboxChecked == '0':
                # Special move to element for Firefox
                if 'firefox' in self.driver.capabilities['browserName']:
                    self.scrollFirefox(attendeeCboxParent)
                actions = ActionChains(self.driver)
                actions.move_to_element(attendeeCboxParent)
                actions.pause(1)
                actions.click(attendeeCboxParent)
                actions.perform()
            else:
                print('Attendee checkbox already checked')
        print(f'End track attendance ({attendeeCount}, {len(self.attendees)})')

##################################################################################
def main():
    # Initialize from command line and set up logging
    tcla = TLCAttendance()

    # Read configuration from file
    tcla.readConfig()

    count = tcla.loadAttendanceData()
    if count == 0:
        print("ERROR: No data found in spreadsheet.")
        return

    # Login to TLC
    if (tcla.login()):
        time.sleep(5)

        # Read in attendance data
        tcla.loadAttendanceData()

        # Record attendance
        tcla.recordAttendance()

        time.sleep(5)

        # Logout from TLC
        tcla.logout()

##################################################################################
if __name__ == "__main__":
    main()